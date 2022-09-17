import re
import os
from openpyxl import Workbook
province_name = ['福建', '河北', '河南', '湖南', '湖北', '广东', '北京', '上海', '天津', '重庆', '四川', '黑龙江', '内蒙古', '西藏', '宁夏', '浙江', '山东',
                 '安徽', '江西', '山西', '吉林', '江苏', '云南', '贵州', '陕西', '甘肃', '青海', '辽宁', '广西', '新疆', '海南']


def verify_name(name_str):
    area_name = ''
    for name in province_name:
        if name in name_str:
            area_name = name
            break
    return area_name


def get_Local_data(content):
    area_num = {}
    if re.search(r"[^\u4e00-\u9fa5]本土病例.*。", content) is None:
        area_num['本土病例'] = '0'
        return area_num
    else:
        information = (re.search(r"[^\u4e00-\u9fa5]本土病例.*。", content)).group()

        all_increase = (re.search(r'\d+', information)).group()  # all_increase 为本土新增的病例数
        area_num[(re.search(r'本土病例', information)).group()] = all_increase

        # 获取出每个省份新增人数的信息
        province_increase = (re.search(r'（.*?）', information)).group()

        # 解决文本格式不一致的问题,比如说xx省x例，其中xx市几例；xx省x例，其中xx市x例
        if '；' in province_increase:
            increase_list = province_increase.split('；')
        else:
            increase_list = province_increase.split('，')

        for area_increase in increase_list:
            area_name = re.search(r'[\u4e00-\u9fa5]+', area_increase).group()
            # 解决出现类似：本土新增x例（均在福建）的问题
            if re.search(r'[0-9]+', area_increase) is None:
                area_name = verify_name(area_name)
                area_num[area_name] = all_increase
            else:
                area_name = verify_name(area_name)
                increase_num = re.search(r'[0-9]+', area_increase).group()
                area_num[area_name] = increase_num

        # 测试数据
        for data in area_num:
            print(data, area_num[data])
        return area_num


# 获取由无症状感染者变为确诊患者的数据
def change_to_confirmed(content):
    # 创建字典，key的值是某个省份，value的值是对应省份的由无症状感染者变为确诊患者的数量
    change_dict = {}
    if re.search(r"[^\u4e00-\u9fa5]本土病例.*。", content) is None:
        change_dict["由无症状感染者变为确诊病例"] = '0'
        return change_dict
    else:
        information = (re.search(r"[^\u4e00-\u9fa5]本土病例.*。", content)).group()
        # 提取相关文本的内容
        if re.search(r'含[0-9]+例由无症状感染者.*?。', information) is None:
            change_dict["由无症状感染者变为确诊病例"] = '0'
            return change_dict
        else:
            change_info = re.search(r'含[0-9]+例由无症状感染者.*?。', information).group()
            # 提取由无症状感染者变为确诊患者的数量
            change_num = re.search(r'[0-9]+', change_info).group()
            change_dict['由无症状感染者转为确诊病例'] = change_num
            # 提取无症状感染者变为确诊病例的患者所在的省份
            change_cases_province_info = re.search(r'（.*?）', change_info).group()
            change_list = change_cases_province_info.split('，')
            # print(change_list)
            for change in change_list:
                area_name = re.search(r'[\u4e00-\u9fa5]+', change).group()
                if re.search(r'[0-9]+', change) is None:
                    area_name = verify_name(area_name)
                    change_dict[area_name] = change_num
                else:
                    area_name = verify_name(area_name)
                    increase_num = re.search(r'[0-9]+', change).group()
                    change_dict[area_name] = increase_num
            # for data in change_dict:
            #     print(data, change_dict[data])
            return change_dict


def get_Asymptomatic_infected_data(content):
    asymptomatic_infected_dict = {}
    if re.search(r'新增无症状感染者[0-9]+例.*。', content) is None:
        asymptomatic_infected_dict['新增无症状感染者'] = '0'
        return asymptomatic_infected_dict
    else:
        asymptomatic_infected_info = (re.search(r"新增无症状感染者[0-9]+例.*。", content)).group()
        if re.search(r'本土[0-9]+例（.*?）', asymptomatic_infected_info) is None:
            asymptomatic_infected_dict['新增无症状感染者'] = 0
            return asymptomatic_infected_dict
        else:
            mainland_info = re.search(r'本土[0-9]+例（.*?）', asymptomatic_infected_info).group()
            mainland_asymptomatic_infected_increase = re.search(r'[0-9]+', mainland_info).group()
            asymptomatic_infected_dict['新增无症状感染者'] = int(mainland_asymptomatic_infected_increase)
            province_asymptomatic_infected_increase = re.search(r'（.*?）', mainland_info).group()
            if len(province_asymptomatic_infected_increase.split('；')) == 1:
                province_list = province_asymptomatic_infected_increase.split('，')
            else:
                province_list = province_asymptomatic_infected_increase.split('；')
            for province in province_list:
                area_name = re.search(r'[\u4e00-\u9fa5]+', province).group()
                if re.search(r'[0-9]+', province) is None:
                    area_name = verify_name(area_name)
                    asymptomatic_infected_dict[area_name] = mainland_asymptomatic_infected_increase
                else:
                    area_name = verify_name(area_name)
                    increase_num = re.search(r'[0-9]+', province).group()
                    asymptomatic_infected_dict[area_name] = increase_num
            return asymptomatic_infected_dict


def refine_dict(dictionary):
    for province in province_name:
        if province not in dictionary:
            dictionary[province] = 0
    return dictionary


def save_to_excel(local_key, local_value, change_key, change_value, asymptomatic_key, asymptomatic_value, filename):
    # 创建一个新的Excel
    wb = Workbook()  # 创建Excel档案
    ws = wb.active  # 使用预设的工作表
    ws.title = '每日本土新增'
    # 写入数据
    local_low = len(local_key)
    for row in range(1, local_low+1):
        ws.cell(row=row, column=1, value=local_key[row-1])
        ws.cell(row=row, column=2, value=int(local_value[row-1]))
    ws1 = wb.create_sheet("由无症状感染者变为确诊病例")
    for row in range(1, len(change_key)+1):
        ws1.cell(row=row, column=1, value=change_key[row-1])
        ws1.cell(row=row, column=2, value=int(change_value[row-1]))
    # 保存
    ws2 = wb.create_sheet("本土新增无症状感染者")
    for row in range(1, len(asymptomatic_key)+1):
        ws2.cell(row=row, column=1, value=asymptomatic_key[row-1])
        ws2.cell(row=row, column=2, value=int(asymptomatic_value[row-1]))
    wb.save(f'C:/Users/OSHMK/Desktop/wjwExcel/{filename}.xlsx')


if __name__ == '__main__':
    path = 'C:/Users/OSHMK/Desktop/wjwData'
    dirs = os.listdir(path)
    dirs.reverse()
    # print(dirs)
    for file in dirs:
        if file == '2021-05-15-截至5月14日24时新型冠状病毒肺炎疫情最新情况.txt':
            break
        f = open(file=f'C:/Users/OSHMK/Desktop/wjwData/{file}', mode='r', encoding='utf-8')
        text = f.read()
        local_data = refine_dict(get_Local_data(text))
        change_data = refine_dict(change_to_confirmed(text))
        asymptomatic_data = refine_dict(get_Asymptomatic_infected_data(text))

        local_data_key = list(local_data.keys())
        local_data_value = list(local_data.values())

        change_data_key = list(change_data.keys())
        change_data_value = list(change_data.values())

        asymptomatic_data_key = list(asymptomatic_data.keys())
        asymptomatic_data_value = list(asymptomatic_data.values())

        save_to_excel(local_data_key, local_data_value, change_data_key, change_data_value, asymptomatic_data_key, asymptomatic_data_value, file)
        f.close()
        print(f"{file}提取完成")
    print("全部提取完毕")
